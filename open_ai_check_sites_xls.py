import socket
import ssl
import requests
import whois
import dns.resolver
from urllib.parse import urlparse
from datetime import datetime

from openpyxl import load_workbook

def analyse_site(url):

    resultat = {
        "accessibilite": "Inconnu",
        "certificat_ssl": "Inconnu",
        "hebergement": "Inconnu",
        "date_mise_en_ligne": "Inconnue",
        "date_maj": "Inconnue",
        "dnssec": "Non"
    }

    try:

        if not url.startswith(("http://", "https://")):
            url = "https://" + url

        parsed = urlparse(url)
        domaine = parsed.netloc

        # ==================================================
        # ACCESSIBILITE
        # ==================================================

        try:
            response = requests.get(
                url,
                timeout=15,
                allow_redirects=True
            )

            if response.status_code == 200:
                resultat["accessibilite"] = "Accessible"
            else:
                resultat["accessibilite"] = f"HTTP {response.status_code}"

            last_mod = response.headers.get("Last-Modified")

            if last_mod:
                resultat["date_maj"] = last_mod

        except Exception:
            resultat["accessibilite"] = "Inaccessible"

        # ==================================================
        # SSL
        # ==================================================

        try:

            context = ssl.create_default_context()

            with socket.create_connection(
                (domaine, 443),
                timeout=10
            ) as sock:

                with context.wrap_socket(
                    sock,
                    server_hostname=domaine
                ) as ssock:

                    cert = ssock.getpeercert()

                    expiration = datetime.strptime(
                        cert["notAfter"],
                        "%b %d %H:%M:%S %Y %Z"
                    )

                    jours = (expiration - datetime.utcnow()).days

                    resultat["certificat_ssl"] = (
                        f"Valide ({jours} jours)"
                    )

        except Exception:
            resultat["certificat_ssl"] = "Non valide"

        # ==================================================
        # HEBERGEMENT
        # ==================================================

        try:

            ip = socket.gethostbyname(domaine)

            resultat["hebergement"] = ip

        except Exception:
            pass

        # ==================================================
        # DATE CREATION DOMAINE
        # ==================================================

        try:

            info = whois.whois(domaine)

            creation = info.creation_date

            if isinstance(creation, list):
                creation = creation[0]

            if creation:
                resultat["date_mise_en_ligne"] = creation.strftime(
                    "%d/%m/%Y"
                )

        except Exception:
            pass

        # ==================================================
        # DNSSEC
        # ==================================================

        try:

            dns.resolver.resolve(
                domaine,
                "DNSKEY"
            )

            resultat["dnssec"] = "Oui"

        except Exception:
            resultat["dnssec"] = "Non"

    except Exception:
        pass

    return resultat

def traiter_excel(fichier_excel):

    wb = load_workbook(fichier_excel)

    ws = wb.active

    for row in range(2, ws.max_row + 1):

        numero = ws[f"A{row}"].value
        site = ws[f"B{row}"].value

        # Ignorer les lignes sans numéro d'ordre

        if not numero:
            continue

        if not site:
            continue

        print(f"Analyse : {site}")

        res = analyse_site(str(site))

        ws[f"E{row}"] = res["accessibilite"]
        ws[f"F{row}"] = res["certificat_ssl"]
        ws[f"G{row}"] = res["hebergement"]
        ws[f"H{row}"] = res["date_mise_en_ligne"]
        ws[f"I{row}"] = res["date_maj"]
        ws[f"J{row}"] = res["dnssec"]

    wb.save("Resultat_Audit_Sites_open_ai.xlsx")

    print("Traitement terminé.")

if __name__ == "__main__":

    #fichier = "sites.xlsx"
    #fichier = "sites_equipe_4.xlsx"
    fichier = "sites_20.xlsx"

    traiter_excel(fichier)