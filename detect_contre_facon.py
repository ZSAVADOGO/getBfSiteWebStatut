import datetime
import socket
import ssl
import sys
import time
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import lru_cache
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import dns.resolver
import whois
from ipwhois import IPWhois
import urllib3

# Désactivation des alertes de sécurité SSL dans la console
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Session HTTP globale simulant un navigateur
HTTP_SESSION = requests.Session()
HTTP_SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Cache-Control": "no-cache",
    "Connection": "keep-alive"
})

# Dictionnaire de correspondance pour catégoriser automatiquement selon le sous-domaine
SECTEURS_MAPPING = {
    "commerce": "COMMERCE / INDUSTRIE",
    "industry": "COMMERCE / INDUSTRIE",
    "communication": "COMMUNICATION / MÉDIAS",
    "mediateur": "COMMUNICATION / MÉDIAS",
    "integration": "COOPÉRATION / INTÉGRATION",
    "cooperation": "COOPÉRATION / INTÉGRATION",
    "culture": "CULTURE / TOURISME",
    "tourisme": "CULTURE / TOURISME",
    "agropastoral": "AGRICULTURE / ÉLEVAGE",
    "sante": "SANTE / HYGIÈNE",
    "finances": "ÉCONOMIE / FINANCES",
    "justice": "JUSTICE / DROITS HUMAINS",
    "education": "ÉDUCATION / RECHERCHE",
}

@lru_cache(maxsize=1024)
def cached_ip_lookup(ip):
    try:
        obj = IPWhois(ip, timeout=3.0)
        rdap_res = obj.lookup_rdap(depth=1)
        asn_desc = rdap_res.get("asn_description", "").upper()
        if "OVH" in asn_desc: return "OVH"
        elif "AMAZON" in asn_desc or "AWS" in asn_desc: return "AWS"
        elif "MICROSOFT" in asn_desc or "AZURE" in asn_desc: return "Azure"
        elif "CLOUDFLARE" in asn_desc: return "Cloudflare"
        elif "GOOGLE" in asn_desc: return "Google Cloud"
        elif "ONATEL" in asn_desc: return "ONATEL"
        elif "ANPTIC" in asn_desc: return "ANPTIC"
        elif asn_desc: return asn_desc.split(',')[0].split('-')[0].strip()
    except Exception:
        pass
    return "Inaccessible"

@lru_cache(maxsize=1024)
def cached_whois_lookup(domain):
    try:
        info = whois.whois(domain)
        creation = info.creation_date
        if isinstance(creation, list):
            creation = creation[0]
        if creation:
            return creation
    except Exception:
        pass
    return None

def determiner_secteur(domain):
    domain_lower = domain.lower()
    for cle, secteur in SECTEURS_MAPPING.items():
        if cle in domain_lower:
            return secteur
    return "AUTRES INSTITUTIONS / ADMINISTRATIONS"

def parse_http_date(date_str):
    if not date_str or "Inconnue" in date_str:
        return "Inconnue"
    months = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
    }
    try:
        parts = date_str.strip().split()
        if len(parts) >= 5:
            day = int(parts[0].replace(",", ""))
            month = months[parts[1]]
            year = int(parts[2])
            time_parts = parts[3].split(":")
            dt = datetime.datetime(year, month, day, int(time_parts[0]), int(time_parts[1]), int(time_parts[2]))
            return dt.strftime("%d/%m/%Y %Hh-%M-%S")
    except Exception:
        pass
    return "Inconnue"

def get_http_status_explanation(status_code):
    explanations = {
        "200": "OK - Site fonctionnel.",
        "301": "Redirection permanente.", "302": "Redirection temporaire.",
        "400": "Requête incorrecte.", "401": "Non autorisé.", 
        "403": "Interdit (Accès bloqué / Pare-feu).",
        "404": "Page non trouvée.", "500": "Erreur interne du serveur.", 
        "502": "Mauvaise passerelle.", "503": "Service indisponible.", "504": "Passerelle expirée."
    }
    return explanations.get(str(status_code), f"Code {status_code} - Statut alternatif.")

""" def scan_single_domain_susceptibility(url):
    if not url.startswith("http://") and not url.startswith("https://"):
        url_to_request = "https://" + url
    else:
        url_to_request = url

    parsed_url = urlparse(url_to_request)
    domain = parsed_url.netloc if parsed_url.netloc else url
    secteur = determiner_secteur(domain)
    scan_time = datetime.datetime.now().strftime("%d/%m/%Y %Hh-%M-%S")
    
    results = {
        "url": url, "secteur": secteur, "disponible": "NON", "statut_http": "Inconnu", 
        "certificat": "INCONNU", "hosting": "non défini", "date_mise_en_ligne": "Non définie", 
        "last_modified": "Inconnue", "dnssec": "Non activé", "observation": "", "date_scan": scan_time, 
        "classification": "Douteux"
    }

    # 1. Infrastructure réseau
    try:
        socket.setdefaulttimeout(3.0)
        ip = socket.gethostbyname(domain)
        results["hosting"] = cached_ip_lookup(ip)
    except Exception:
        results["hosting"] = "Inaccessible (DNS)"

    try:
        resolver = dns.resolver.Resolver()
        resolver.timeout = 2.0
        resolver.lifetime = 2.0
        answers = resolver.resolve(domain, 'RRSIG', raise_on_no_answer=True)
        if len(answers) > 0:
            results["dnssec"] = "Activé"
    except Exception:
        results["dnssec"] = "Non activé"

    creation_date = cached_whois_lookup(domain)

    if creation_date:
        # CORRECTION BUG 1 : Nettoyage de la timezone WHOIS pour éviter le crash de soustraction
        if hasattr(creation_date, 'tzinfo') and creation_date.tzinfo is not None:
            creation_date = creation_date.replace(tzinfo=None)
        results["date_mise_en_ligne"] = creation_date.strftime("%d/%m/%Y")

    # 2. Détection Contrefaçon
    est_officiel_gov = domain.endswith(".gov.bf")
    est_recent = False
    if creation_date:
        age_jours = (datetime.datetime.now() - creation_date).days
        if age_jours < 90:
            est_recent = True

    if not est_officiel_gov and ("gov" in domain or "sante" in domain or "gouv" in domain or "faso" in domain):
        results["classification"] = "Contrefaçon"
        results["observation"] = "CRITIQUE : Mots-clés étatiques sans l'extension officielle .gov.bf."
    elif est_recent:
        results["classification"] = "Contrefaçon"
        results["observation"] = "SUSPECT : Domaine créé il y a moins de 90 jours."

    # 3. Requête HTTP
    if results["classification"] != "Contrefaçon":
        try:
            response = HTTP_SESSION.get(url_to_request, timeout=(3.0, 4.0), allow_redirects=True, verify=True)
            results["statut_http"] = str(response.status_code)
            results["disponible"] = "OUI (Fonctionnel)" if response.status_code == 200 else f"OUI ({response.status_code})"
            
            html_content = response.text.lower()
            is_apache_default = "apache2 default page" in html_content or "it works!" in html_content
            is_http = response.url.startswith("http://")

            if is_apache_default:
                results["classification"] = "Douteux"
                results["observation"] = "ALERTE : Page d'installation par défaut Apache (It works!)."
            elif is_http:
                results["classification"] = "Douteux"
                results["observation"] = "ATTENTION : Le site transite uniquement en HTTP (Non sécurisé)."
            elif response.status_code == 200:
                results["classification"] = "Sûr"
                results["observation"] = "Site fonctionnel | Flux HTTPS sécurisé."
            else:
                results["classification"] = "Douteux"
                results["observation"] = get_http_status_explanation(response.status_code)

            last_mod_raw = response.headers.get("Last-Modified")
            results["last_modified"] = parse_http_date(last_mod_raw) if last_mod_raw else "Inconnue"

        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError):
            try:
                fallback_url = url_to_request.replace("https://", "http://")
                response = HTTP_SESSION.get(fallback_url, timeout=(3.0, 4.0), allow_redirects=True, verify=False)
                results["statut_http"] = str(response.status_code)
                results["disponible"] = "OUI (HTTP)"
                results["classification"] = "Douteux"
                
                if "apache2 default page" in response.text.lower() or "it works!" in response.text.lower():
                    results["observation"] = "ALERTE : Page par défaut Apache (HTTP uniquement)."
                else:
                    results["observation"] = "ATTENTION : Échec HTTPS. Connexion repliée en HTTP."
                
                last_mod_raw = response.headers.get("Last-Modified")
                results["last_modified"] = parse_http_date(last_mod_raw) if last_mod_raw else "Inconnue"
            except Exception:
                results["classification"] = "Douteux"
                results["statut_http"] = "Error"
                results["observation"] = "Le site refuse le HTTPS et le HTTP."
        except Exception:
            results["classification"] = "Douteux"
            results["statut_http"] = "Timeout"
            results["observation"] = "Le serveur ne répond pas (Timeout)."

    # 4. Certificat SSL fine
    if results["classification"] == "Sûr" or (results["statut_http"] != "Inconnu" and not url_to_request.startswith("http://")):
        try:
            context = ssl.create_default_context()
            with socket.create_connection((domain, 443), timeout=2.0) as sock:
                with context.wrap_socket(sock, server_hostname=domain) as ssock:
                    cert = ssock.getpeercert()
                    expire_date = datetime.datetime.strptime(cert.get("notAfter"), "%b %d %H:%M:%S %Y %Z")
                    jours = (expire_date - datetime.datetime.utcnow()).days
                    results["certificat"] = f"{ssock.version()} ({jours} j restants)"
                    if jours <= 0:results["classification"] = "Douteux"
                    results["observation"] = "ALERTE : Le certificat SSL a expiré."
        except Exception:results["certificat"] = "INCONNU / ERREUR SSL"
        if results["classification"] == "Sûr":
            results["classification"] = "Douteux"
            results["observation"] = "ATTENTION : Erreur d'établissement SSL."
        else:results["certificat"] = "NON APPLICABLE"
        return results """

def scan_single_domain_susceptibility(url):
    if not url.startswith("http://") and not url.startswith("https://"):
        url_to_request = "https://" + url
    else:
        url_to_request = url

    parsed_url = urlparse(url_to_request)
    domain = parsed_url.netloc if parsed_url.netloc else url
    secteur = determiner_secteur(domain)
    scan_time = datetime.datetime.now().strftime("%d/%m/%Y %Hh-%M-%S")
    
    results = {
        "url": url, "secteur": secteur, "disponible": "NON", "statut_http": "Inconnu", 
        "certificat": "INCONNU", "hosting": "non défini", "date_mise_en_ligne": "Non définie", 
        "last_modified": "Inconnue", "dnssec": "Non activé", "observation": "", "date_scan": scan_time, 
        "classification": "Douteux"
    }

    # 1. Infrastructure réseau
    try:
        socket.setdefaulttimeout(3.0)
        ip = socket.gethostbyname(domain)
        results["hosting"] = cached_ip_lookup(ip)
    except Exception:
        results["hosting"] = "Inaccessible (DNS)"

    try:
        resolver = dns.resolver.Resolver()
        resolver.timeout = 2.0
        resolver.lifetime = 2.0
        answers = resolver.resolve(domain, 'RRSIG', raise_on_no_answer=True)
        if len(answers) > 0:
            results["dnssec"] = "Activé"
    except Exception:
        results["dnssec"] = "Non activé"

    creation_date = cached_whois_lookup(domain)
    if creation_date:
        if hasattr(creation_date, 'tzinfo') and creation_date.tzinfo is not None:
            creation_date = creation_date.replace(tzinfo=None)
        results["date_mise_en_ligne"] = creation_date.strftime("%d/%m/%Y")

    # 2. OPTIMISATION : Détection Contrefaçon adaptée (.bf et .gov.bf légitimes)
    est_officiel_gov = domain.endswith(".gov.bf")
    est_officiel_bf = domain.endswith(".bf")
    est_recent = False
    
    if creation_date:
        now_naive = datetime.datetime.now()
        age_jours = (now_naive - creation_date).days
        if age_jours < 90:
            est_recent = True

    mots_cles_etatiques = ["gov", "sante", "gouv", "faso", "ministere", "institution"]
    contient_mot_cle = any(mot in domain.lower() for mot in mots_cles_etatiques)

    # Si le domaine usurpe un nom de l'état sans être en .bf ou .gov.bf
    if contient_mot_cle and not est_officiel_bf:
        results["classification"] = "Contrefaçon"
        results["observation"] = "CRITIQUE : Mots-clés étatiques repérés sur une extension non nationale."
    # Si le domaine prétend être gouvernemental mais est extrêmement récent
    elif est_recent and (est_officiel_bf or est_officiel_gov):
        results["classification"] = "Contrefaçon"
        results["observation"] = "SUSPECT : Domaine national créé il y a moins de 90 jours."
    # Si le domaine utilise un mot-clé étatique fort et s'arrête au .bf au lieu du .gov.bf
    elif contient_mot_cle and est_officiel_bf and not est_officiel_gov:
        results["classification"] = "Douteux"
        results["observation"] = "ATTENTION : Site institutionnel utilisant l'extension .bf simple au lieu de .gov.bf."

    # 3. Requête HTTP
    if results["classification"] != "Contrefaçon":
        try:
            response = HTTP_SESSION.get(url_to_request, timeout=(3.0, 4.0), allow_redirects=True, verify=True)
            results["statut_http"] = str(response.status_code)
            results["disponible"] = "OUI (Fonctionnel)" if response.status_code == 200 else f"OUI ({response.status_code})"
            
            html_content = response.text.lower()
            is_apache_default = "apache2 default page" in html_content or "it works!" in html_content
            is_http = response.url.startswith("http://")

            if is_apache_default:
                results["classification"] = "Douteux"
                results["observation"] += " | ALERTE : Page d'installation par défaut Apache (It works!)." if results["observation"] else "ALERTE : Page d'installation par défaut Apache (It works!)."
            elif is_http:
                results["classification"] = "Douteux"
                results["observation"] += " | ATTENTION : Le site transite uniquement en HTTP (Non sécurisé)." if results["observation"] else "ATTENTION : Le site transite uniquement en HTTP (Non sécurisé)."
            elif response.status_code == 200:
                # Si aucune anomalie de domaine n'a été détectée en amont, le site est Sûr
                if not results["observation"]:
                    results["classification"] = "Sûr"
                    results["observation"] = "Site fonctionnel | Flux HTTPS sécurisé."
            else:
                results["classification"] = "Douteux"
                explication = get_http_status_explanation(response.status_code)
                results["observation"] += f" | {explication}" if results["observation"] else explication

            last_mod_raw = response.headers.get("Last-Modified")
            results["last_modified"] = parse_http_date(last_mod_raw) if last_mod_raw else "Inconnue"

        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError):
            try:
                fallback_url = url_to_request.replace("https://", "http://")
                response = HTTP_SESSION.get(fallback_url, timeout=(3.0, 4.0), allow_redirects=True, verify=False)
                results["statut_http"] = str(response.status_code)
                results["disponible"] = "OUI (HTTP)"
                results["classification"] = "Douteux"

                msg = "ALERTE : Page par défaut Apache (HTTP uniquement)." if "apache2 default page" in response.text.lower() or "it works!" in response.text.lower() else "ATTENTION : Échec HTTPS. Connexion repliée en HTTP."
                results["observation"] = f"{results['observation']} | {msg}" if results["observation"] else msg
                last_mod_raw = response.headers.get("Last-Modified")
                results["last_modified"] = parse_http_date(last_mod_raw) if last_mod_raw else "Inconnue"
            except Exception:
                results["classification"] = "Douteux"
                results["statut_http"] = "Error"
                results["observation"] = "Le site refuse le HTTPS et le HTTP."
        except Exception:
            results["classification"] = "Douteux"
            results["statut_http"] = "Timeout"
            results["observation"] = "Le serveur ne répond pas (Timeout)."
    # 4. Certificat SS
    if results["classification"] == "Sûr" or (results["statut_http"] != "Inconnu" and not url_to_request.startswith("http://")):
        try:
            context = ssl.create_default_context()
            with socket.create_connection((domain, 443), timeout=2.0) as sock:
                with context.wrap_socket(sock, server_hostname=domain) as ssock:
                    cert = ssock.getpeercert()
                    expire_date = datetime.datetime.strptime(cert.get("notAfter"), "%b %d %H:%M:%S %Y %Z")
                    if expire_date.tzinfo is not None:
                        expire_date = expire_date.replace(tzinfo=None)
                    now_utc_naive = datetime.datetime.utcnow()
                    jours = (expire_date - now_utc_naive).days
                    results["certificat"] = f"{ssock.version()} ({jours} j restants)"
                    if jours <= 0:
                        results["classification"] = "Douteux"
                        results["observation"] = "ALERTE : Le certificat SSL a expiré."
        except Exception:
            results["certificat"] = "INCONNU / ERREUR SSL"
            if results["classification"] == "Sûr":
                results["classification"] = "Douteux"
                results["observation"] = "ATTENTION : Erreur d'établissement SSL."
            else:
                results["certificat"] = "NON APPLICABLE"
    return results



def process_excel_concurrent(filepath: str, max_workers: int = 30):
    start_time = time.time()
    
    # 1. Lecture de la Feuille 2 du fichier d'origine
    wb_source = openpyxl.load_workbook(filepath, data_only=True)
    if len(wb_source.sheetnames) < 2:
        print("❌ Erreur : Le fichier ne contient pas de deuxième feuille.")
        return
    
    # Sélection de la deuxième feuille (index 1)
    ws_source = wb_source.worksheets[1]
    
    tasks = []
    rows = list(ws_source.iter_rows(min_row=1))
    
    # 2. Collecte et filtrage des URL à analyser
    for row_idx, row in enumerate(rows, start=1):
        col_a = row[0].value
        col_b = row[1].value if len(row) > 1 else None

        a_est_nombre = isinstance(col_a, (int, float)) and not isinstance(col_a, bool)
        b_est_chaine = isinstance(col_b, str) and col_b.strip() != ""

        if not a_est_nombre or not b_est_chaine:
            if col_a is not None or col_b is not None:
                print(f"  [Ligne {row_idx}] Ignorée — A={repr(col_a)} | B={repr(str(col_b)[:60])}")
            continue

        tasks.append((row_idx, col_b.strip()))

    total = len(tasks)
    if total == 0:
        print("❌ Aucune URL valide trouvée à traiter sur la feuille 2.")
        return
        
    print(f"\n🚀 Lancement de l'analyse sur {total} sites ({max_workers} threads)\n")

    # 3. Exécution des requêtes concurrentes (I/O Bound)
    results = []
    completed = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(scan_single_domain_susceptibility, url): (row_idx, url)
            for row_idx, url in tasks
        }

        for future in as_completed(future_map):
            row_idx, url = future_map[future]
            completed += 1
            try:
                data = future.result()
                if data:
                    # On garde l'URL dans les données pour l'écriture finale
                    data["url_analyse"] = url
                    results.append(data)
                    print(f"[{completed:>4}/{total}] {url} → HTTP {data['statut_http']} → {data.get('observation', 'Analysé')}")
                    print(f"  Disponible : {data['disponible']}")
                    print(f"  SSL/TLS    : {data['certificat']}")
                    print(f"  Hébergeur  : {data['hosting']}")
                    print(f"  Création   : {data['date_mise_en_ligne']}")
                    print(f"  Modif.     : {data['last_modified']}")
                    print(f"  DNSSEC     : {data['dnssec']}")
                    print(f"  Observation: {data['observation']}")
                    print("─" * 76)
            
            except Exception as exc:
                print(f"  ❌ ERREUR ligne {row_idx} ({url}) : {exc}")

    # 4. Création du nouveau classeur avec les feuilles catégorisées
    wb_new = openpyxl.Workbook()
    
    # Suppression de la feuille par défaut et création des 3 catégories
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)
    
    ws_surs = wb_new.create_sheet(title="Sites Sûrs")
    ws_douteux = wb_new.create_sheet(title="Sites Douteux")
    ws_contrefacons = wb_new.create_sheet(title="Contrefaçons")

    # En-têtes pour les nouvelles feuilles
    headers = [
        "URL", "Disponible", "Certificat", "Hébergeur", 
        "Date Mise en Ligne", "Last Modified", "DNSSEC", "Observation", "Date Scan"
    ]
    
    for ws in [ws_surs, ws_douteux, ws_contrefacons]:
        ws.append(headers)

    # 5. Tri et répartition des résultats
    for data in results:
        # Récupération propre des valeurs
        row_data = [
            data.get("url_analyse"),
            data.get("disponible"),
            data.get("certificat"),
            data.get("hosting"),
            data.get("date_mise_en_ligne"),
            data.get("last_modified"),
            data.get("dnssec"),
            data.get("observation", ""),
            data.get("date_scan")
        ]
        
        # Normalisation du texte de l'observation pour le filtrage (minuscules, sans espaces superflus)
        obs = str(data.get("observation", "")).lower().strip()
        
        # Dispatching selon le contenu de la colonne "observation"
        if "contrefacon" in obs or "contrefaçon" in obs:
            ws_contrefacons.append(row_data)
        elif "douteux" in obs or "suspect" in obs or "attention" in obs:
            ws_douteux.append(row_data)
        else:
            # Par défaut, si aucune anomalie critique n'est mentionnée
            ws_surs.append(row_data)


    end_time = time.time()
    elapsed_time = end_time - start_time
    minutes = int(elapsed_time // 60)
    seconds = int(elapsed_time % 60)


    horodatage      = datetime.datetime.now().strftime("%d-%m-%Y_%Hh-%M-%S")
    output_filename = f"audit_contrefacon_domaines_{horodatage}.xlsx"
    wb_new.save(output_filename)


    print(f"\n😍😍 Audit terminé avec succès ! Fichier généré => {output_filename}")
    print(f"Temps de traitement global : {minutes:02d}mm/{seconds:02d}ss")


if __name__ == "__main__":
    fichier_liste = sys.argv[1] if len(sys.argv) > 1 else "plateformes_burkina_faso.xlsx"
    #fichier_liste = sys.argv[1] if len(sys.argv) > 1 else "zak_924_a_972_plateformes_burkina_faso.xlsx"
    #fichier_liste = sys.argv[1] if len(sys.argv) > 1 else "rayhana_973_a_1021_plateformes_burkina_faso.xlsx"
    #fichier_liste = sys.argv[1] if len(sys.argv) > 1 else "equipe_e1_sites_sawadogo_idrissa_071838.xlsx"

    workers = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    process_excel_concurrent(fichier_liste)


