import datetime
import socket
import ssl
import sys
import time
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import lru_cache
import requests
import openpyxl
import dns.resolver
import whois
from ipwhois import IPWhois

# 1. OPTIMISATION : Session requests globale et partagée avec des en-têtes avancés (Simule un navigateur)
HTTP_SESSION = requests.Session()
HTTP_SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Cache-Control": "no-cache",
    "Connection": "keep-alive"
})

# 2. OPTIMISATION : Cache LRU pour IPWhois/RDAP (Même IP -> Résultat réutilisé)
@lru_cache(maxsize=1024)
def cached_ip_lookup(ip):
    try:
        obj = IPWhois(ip, timeout=3.0)
        rdap_res = obj.lookup_rdap(depth=1)
        asn_desc = rdap_res.get("asn_description", "").upper()
        
        if "OVH" in asn_desc: return "OVH"
        elif "AMAZON" in asn_desc or "AWS" in asn_desc: return "AWS"
        elif "MICROSOFT" in asn_desc or "AZURE" in asn_desc: return "Azure"
        elif "CLOUDFLARE" in asn_desc: return "Cloudflare"
        elif "GOOGLE" in asn_desc: return "Google Cloud"
        elif "OFFICE NATIONAL DES TELECOMMUNICATIONS" in asn_desc or "ONATEL" in asn_desc: return "ONATEL"
        elif "AGENCE NATIONALE DE PROMOTION DES TIC" in asn_desc or "ANPTIC" in asn_desc: return "ANPTIC"
        elif asn_desc: return asn_desc.split(',')[0].split('-')[0].strip()
    except Exception:
        pass
    return "Inaccessible (Timeout)"

# 3. OPTIMISATION : Cache LRU pour WHOIS (Même domaine -> Résultat réutilisé)
@lru_cache(maxsize=1024)
def cached_whois_lookup(domain):
    try:
        info = whois.whois(domain)
        creation = info.creation_date
        if isinstance(creation, list):
            creation = creation[0]
        if creation:
            return creation.strftime("%d/%m/%Y")
    except Exception:
        pass
    return "Masqué / Erreur"

def parse_http_date(date_str):
    """Convertit la date HTTP au format 'dd/mm/yyyy Hh-MM-SS'."""
    if not date_str or "Inconnue" in date_str:
        return "Inconnue"
    months = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
    }
    try:
        parts = date_str.strip().split()
        if len(parts) >= 5:
            day = int(parts[0].replace(",", ""))
            month = months[parts[1]]
            year = int(parts[2])
            time_parts = parts[3].split(":")
            dt = datetime.datetime(year, month, day, int(time_parts[0]), int(time_parts[1]), int(time_parts[2]))
            return dt.strftime("%d/%m/%Y %Hh-%M-%S")
    except Exception:
        pass
    return "Inconnue"

def get_http_status_explanation(status_code):
    explanations = {
        "200": "OK - Site fonctionnel et accessible publiquement.",
        "301": "Redirection permanente.", "302": "Redirection temporaire.",
        "400": "Requête incorrecte.", "401": "Non autorisé.", 
        "403": "Interdit (Accès bloqué / Pare-feu strict ou restriction WAF).",
        "404": "Page non trouvée.", "500": "Erreur interne du serveur.", 
        "502": "Mauvaise passerelle.", "503": "Service indisponible.", "504": "Passerelle expirée."
    }
    return explanations.get(str(status_code), f"Code HTTP {status_code} - Statut personnalisé.")

def analyze_domain_infrastructure(domain):
    """Analyse l'hébergeur, le DNSSEC et la création du domaine avec timeouts calibrés."""
    hosting = "non défini"
    dnssec_status = "Non activé"
    creation_date = "Non définie"
    
    if not domain:
        return hosting, dnssec_status, creation_date

    try:
        socket.setdefaulttimeout(3.0) # 8. OPTIMISATION : Timeouts réduits
        ip = socket.gethostbyname(domain)
        hosting = cached_ip_lookup(ip)
    except Exception:
        hosting = "Inaccessible (DNS Error)"

    try:
        resolver = dns.resolver.Resolver()
        resolver.timeout = 2.0
        resolver.lifetime = 2.0
        answers = resolver.resolve(domain, 'RRSIG', raise_on_no_answer=True)
        if len(answers) > 0:
            dnssec_status = "Activé"
    except Exception:
        dnssec_status = "Non activé"

    creation_date = cached_whois_lookup(domain)
    return hosting, dnssec_status, creation_date

def check_website_complete(url):
    """Effectue l'analyse complète d'un site web de manière ultra-rapide."""
    if not url.startswith("http://") and not url.startswith("https://"):
        url_to_request = "https://" + url
    else:
        url_to_request = url

    parsed_url = urlparse(url_to_request)
    domain = parsed_url.netloc
    scan_time = datetime.datetime.now().strftime("%d/%m/%Y %Hh-%M-%S")
    
    results = {
        "url": url, "disponible": "NON", "statut_http": "Inconnu", "certificat": "INCONNU", 
        "hosting": "non défini", "date_mise_en_ligne": "Non définie", "last_modified": "Inconnue", 
        "dnssec": "Non activé", "observation": "", "date_scan": scan_time, "is_http_only": False
    }

    hosting, dnssec, creation = analyze_domain_infrastructure(domain)
    results["hosting"] = hosting
    results["dnssec"] = dnssec
    results["date_mise_en_ligne"] = creation

    # 8. OPTIMISATION : Timeouts réseau réduits/calibrés (3s connexion, 4s lecture)
    try:
        response = HTTP_SESSION.get(url_to_request, timeout=(3.0, 4.0), allow_redirects=True, verify=True)
        results["statut_http"] = str(response.status_code)
        
        if response.url.startswith("http://"):
            results["is_http_only"] = True

        if response.status_code == 200:
            results["disponible"] = "OUI (Fonctionnel)"
        else:
            results["disponible"] = f"OUI (Statut alternatif : {response.status_code})"

        explanation = get_http_status_explanation(response.status_code)
        
        # Détection du cas spécifique : Page par défaut Apache (It works!)
        content_lower = response.text.lower()
        if "apache2 default page" in content_lower or "it works!" in content_lower:
            explanation += " | ALERTE : Le serveur affiche la page d'installation par défaut Apache (It works!). Le site n'est pas configuré."

        if results["is_http_only"]:
            explanation += " | ALERTE : Le site utilise le protocole HTTP non sécurisé."
        else:
            explanation += " | Flux HTTPS sécurisé."
            
        results["observation"] = explanation
        last_mod_raw = response.headers.get("Last-Modified")
        results["last_modified"] = parse_http_date(last_mod_raw) if last_mod_raw else "Inconnue"
            
    except (requests.exceptions.SSLError, requests.exceptions.ConnectionError):
        # Tentative de repli immédiat en HTTP (Résout le cas de shuri.sante.gov.bf ou pannes SSL)
        try:
            fallback_url = url_to_request.replace("https://", "http://") if url_to_request.startswith("https://") else url_to_request
            response = HTTP_SESSION.get(fallback_url, timeout=(3.0, 4.0), allow_redirects=True, verify=False)
            results["statut_http"] = str(response.status_code)
            results["is_http_only"] = True
            results["disponible"] = "OUI (Accessible en HTTP)"
            
            explanation = get_http_status_explanation(response.status_code)
            content_lower = response.text.lower()
            
            if "apache2 default page" in content_lower or "it works!" in content_lower:
                explanation += " | ALERTE : Serveur accessible uniquement en HTTP et affiche la page par défaut Apache (It works!)."
            else:
                explanation += " | ALERTE : Échec HTTPS. Connexion repliée sur le protocole HTTP non sécurisé."
                
            results["observation"] = explanation
            last_mod_raw = response.headers.get("Last-Modified")
            results["last_modified"] = parse_http_date(last_mod_raw) if last_mod_raw else "Inconnue"
        except Exception:
            results["disponible"] = "NON (Inaccessible)"
            results["statut_http"] = "Erreur Réseau"
            results["observation"] = "Le serveur rejette les connexions HTTPS et ne répond pas aux tentatives HTTP basiques."
            
    except Exception:
        results["disponible"] = "NON (Timeout)"
        results["statut_http"] = "Timeout"
        results["observation"] = "Erreur réseau : Le serveur n'a pas répondu dans le délai imparti."

    # Inspection TLS fine
    if not results["is_http_only"] and results["statut_http"] not in ["Erreur Réseau", "Timeout"]:
        try:
            context = ssl.create_default_context()
            with socket.create_connection((domain, 443), timeout=3.0) as sock:
                with context.wrap_socket(sock, server_hostname=domain) as ssock:
                    tls_version = ssock.version().replace("TLSv", "TLS ")
                    cert = ssock.getpeercert()
                    expire_date = datetime.datetime.strptime(cert.get("notAfter"), "%b %d %H:%M:%S %Y %Z")
                    jours = (expire_date - datetime.datetime.utcnow()).days
                    results["certificat"] = f"{tls_version} ({jours} j restants)"
        except Exception:
            results["certificat"] = "INCONNU"
    else:
        results["certificat"] = "NON APPLICABLE (HTTP)"

    return results

def process_excel_concurrent(filepath, max_workers=30): # 5. OPTIMISATION : Augmenté à 30 workers
    """Parcourt le fichier Excel et écrit les données à l'aide d'un pool asynchrone ultra-rapide."""
    start_time = time.time()

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active 
    tasks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        col_a_val = row[0].value
        col_b_val = row[1].value
        if col_a_val is not None and str(col_a_val).strip() != "" and col_b_val:
            url = str(col_b_val).strip()
            tasks.append((row_idx, url))
    total_sites = len(tasks)
    print(f"\nLancement de l'analyse sur {total_sites} sites")
    completed_count = 0
    # 4. OPTIMISATION : time.sleep(0.5) supprimé pour une vitesse maximale en parallèle
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Soumission globale des tâches en arrière-plan
        futures = {executor.submit(check_website_complete, url): (row_idx, url) for row_idx, url in tasks}
        # 6. OPTIMISATION : as_completed() traite les lignes au fur et à mesure de leur achèvement réel
        for future in as_completed(futures):
            row_idx, url = futures[future]
            completed_count += 1
            try:
                data = future.result()
                # Injection immédiate dans les cellules Excel correspondantes
                ws.cell(row=row_idx, column=5, value=data["disponible"])
                ws.cell(row=row_idx, column=6, value=data["certificat"])
                ws.cell(row=row_idx, column=7, value=data["hosting"])
                ws.cell(row=row_idx, column=8, value=data["date_mise_en_ligne"])
                ws.cell(row=row_idx, column=9, value=data["last_modified"])
                ws.cell(row=row_idx, column=10, value=data["dnssec"])
                ws.cell(row=row_idx, column=11, value=data["observation"])
                ws.cell(row=row_idx, column=12, value=data["date_scan"])
                print(f"N°{completed_count} Analyse du site {url} --> Statut ({data['statut_http']})")
                print(f'"Certificat SSL" = {data["certificat"]}')
                print(f'"Hebergement" = {data["hosting"]}')
                print(f'"Date mise en ligne" = {data["date_mise_en_ligne"]}')
                print(f'"Date dernière modification" = {data["last_modified"]}')
                print(f'"DNSSEC" = {data["dnssec"]}')
                print(f'"Observation" = {data["observation"]}')
                print("------------------------------------------\n")
            except Exception as e:
                print(f"Erreur d'exécution critique sur la Ligne {row_idx} ({url}): {e}")



# 1. Récupérer la date et l'heure actuelles (Format : Jour-Mois-Année_Heures-Minutes)

    horodatage = datetime.datetime.now().strftime("%d-%m-%Y_%Hh-%M-%S")
    output_filename = f"gemini_resultats_sites_{horodatage}.xlsx"
    wb.save(output_filename)

    
    end_time = time.time()
    elapsed_time = end_time - start_time
    minutes = int(elapsed_time // 60)
    seconds = int(elapsed_time % 60)
    
    print(f"😍😍😍 Traitement terminé. Fichier => {output_filename}\n")
    print(f"nTemps mis pour le traitement : {minutes:02d}mm/{seconds:02d}ss\n")

# Exemple d'exécution :
# process_excel_file("vos_sites.xlsx")



if __name__ == "__main__":
    # Nom du fichier par défaut ou passé en argument
    #fichier_liste = sys.argv[1] if len(sys.argv) > 1 else "sites.xlsx"
    fichier_liste = sys.argv[1] if len(sys.argv) > 1 else "sites_equipe_4.xlsx"
    process_excel_concurrent(fichier_liste, max_workers=5)

