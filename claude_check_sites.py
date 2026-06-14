"""
scanner_sites_optimise.py  –  v2
=================================
Corrections v2 (en plus des optimisations de performance v1) :

  FIX A – Faux "NON Accessible" sur sites à certificat auto-signé / CA interne
           → Stratégie de connexion en 3 passes :
               1) HTTPS verify=True  (cas nominal)
               2) HTTPS verify=False (certificat invalide mais site UP)
               3) HTTP              (site sans TLS du tout)

  FIX B – Timeout trop court pour serveurs gouvernementaux lents
           → connect=6 s, read=12 s (au lieu de 3/8)

  FIX C – Détection page Apache par défaut ("It works!")
           → Détection dans le corps HTML + message dédié dans Observation

  FIX D – Session partagée : isolation par thread pour éviter les cookies croisés
           → chaque thread possède sa propre Session (pool maintenu, pas de pollution)

Optimisations v1 maintenues :
  - Détection native de technologies (headers HTTP + HTML)
  - Cache LRU IPWhois/RDAP + whois
  - as_completed() au lieu de l'itération naïve
  - max_workers=30 par défaut
  - time.sleep(0.5) supprimé
"""
import os

import datetime
import socket
import ssl
import sys
import time
import threading
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import urllib3
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
import dns.resolver
import whois
from ipwhois import IPWhois
# Supprime les warnings SSL des certificats invalides (verify=False en passe 2)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ──────────────────────────────────────────────────────────────────────────────
# FIX D – Session par thread (pas de pollution de cookies entre sites)
# ──────────────────────────────────────────────────────────────────────────────
_thread_local = threading.local()

def _get_session() -> requests.Session:
    """Retourne la Session propre au thread courant (créée si absente)."""
    if not hasattr(_thread_local, "session"):
        s = requests.Session()
        retry = Retry(
            total=1,
            backoff_factor=0.2,
            status_forcelist=[502, 503, 504],
            allowed_methods=["GET"],
        )
        adapter = HTTPAdapter(
            max_retries=retry,
            pool_connections=10,
            pool_maxsize=10,
        )
        s.mount("https://", adapter)
        s.mount("http://",  adapter)
        s.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "fr,fr-FR;q=0.8,en-US;q=0.5,en;q=0.3",
            "Cache-Control": "no-cache",
        })
        _thread_local.session = s
    return _thread_local.session


# ──────────────────────────────────────────────────────────────────────────────
# Cache LRU  –  whois et RDAP
# ──────────────────────────────────────────────────────────────────────────────
# Cache thread-safe pour whois et RDAP
# lru_cache est évité ici : il peut mettre en cache des exceptions levées lors
# d'un appel concurrent, et IPWhois n'est pas thread-safe sous lru_cache.
# On utilise un dict + Lock : seul le premier thread fait l'appel réseau,
# les suivants lisent la valeur déjà stockée.
# ──────────────────────────────────────────────────────────────────────────────
_rdap_cache:  dict = {}
_whois_cache: dict = {}
_rdap_lock  = threading.Lock()
_whois_lock = threading.Lock()


def _cached_rdap(ip: str) -> dict:
    if ip not in _rdap_cache:
        with _rdap_lock:
            if ip not in _rdap_cache:          # double-check après acquisition
                try:
                    _rdap_cache[ip] = IPWhois(ip).lookup_rdap(depth=1)
                except Exception:
                    _rdap_cache[ip] = {}       # valeur vide, pas d'erreur propagée
    return _rdap_cache[ip]


def _cached_whois(domain: str):
    if domain not in _whois_cache:
        with _whois_lock:
            if domain not in _whois_cache:
                try:
                    _whois_cache[domain] = whois.whois(domain)
                except Exception:
                    _whois_cache[domain] = None
    return _whois_cache[domain]


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

# Mots-clés de la page Apache par défaut (toutes variantes connues)
_APACHE_DEFAULT_KEYWORDS = [
    "apache2 default page",
    "apache2 ubuntu default page",
    "apache2 debian default page",
    "it works!",
    "this is the default welcome page",
    "apache http server test page",
]


def _is_apache_default_page(html: str) -> bool:
    """Retourne True si la réponse est la page d'accueil Apache par défaut."""
    lower = html.lower()
    return any(kw in lower for kw in _APACHE_DEFAULT_KEYWORDS)


def parse_http_date(date_str: str) -> str:
    if not date_str or "Inconnue" in date_str:
        return "Inconnue"
    months = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
    }
    try:
        parts = date_str.strip().split()
        if len(parts) >= 5:
            day = int(parts[1].replace(",", ""))
            month = months[parts[2]]
            year = int(parts[3])
            h, m, s = (int(x) for x in parts[4].split(":"))
            return datetime.datetime(year, month, day, h, m, s).strftime(
                "%d/%m/%Y %Hh-%M-%S"
            )
    except Exception:
        pass
    return "Inconnue"


def get_http_status_explanation(status_code) -> str:
    """Retourne une explication complète du code HTTP, y compris les codes Cloudflare."""
    explanations = {
        # ── 2xx Succès ───────────────────────────────────────────────────────
        "200": "OK - Site fonctionnel.",
        "201": "Créé - Ressource créée avec succès.",
        "204": "Pas de contenu - Requête traitée sans contenu retourné.",
        "206": "Contenu partiel - Téléchargement partiel en cours.",
        # ── 3xx Redirections ─────────────────────────────────────────────────
        "301": "Redirection permanente - L'URL a changé définitivement.",
        "302": "Redirection temporaire - L'URL est temporairement déplacée.",
        "303": "Voir ailleurs - Redirection vers une autre ressource.",
        "304": "Non modifié - Contenu identique au cache du client.",
        "307": "Redirection temporaire stricte - Méthode HTTP conservée.",
        "308": "Redirection permanente stricte - Méthode HTTP conservée.",
        # ── 4xx Erreurs client ───────────────────────────────────────────────
        "400": "Requête incorrecte - Syntaxe de la requête invalide.",
        "401": "Non autorisé - Authentification requise.",
        "402": "Paiement requis - Accès réservé aux abonnés.",
        "403": "Interdit - Accès bloqué (pare-feu ou restriction IP).",
        "404": "Page non trouvée - La ressource demandée n'existe pas.",
        "405": "Méthode non autorisée - Le serveur refuse cette méthode HTTP.",
        "406": "Non acceptable - Format de réponse incompatible.",
        "407": "Authentification proxy requise.",
        "408": "Timeout de la requête - Le client a mis trop de temps.",
        "409": "Conflit - État incohérent de la ressource.",
        "410": "Supprimé définitivement - La ressource n'existe plus.",
        "429": "Trop de requêtes - Limite de débit atteinte (rate limiting).",
        "451": "Accès refusé pour raisons légales (censure, RGPD…).",
        # ── 5xx Erreurs serveur ──────────────────────────────────────────────
        "500": "Erreur interne du serveur - Anomalie côté serveur.",
        "501": "Non implémenté - Fonctionnalité non supportée par le serveur.",
        "502": "Bad Gateway - Le proxy/reverse proxy a reçu une réponse invalide du serveur amont.",
        "503": "Service indisponible - Serveur en maintenance ou surchargé.",
        "504": "Passerelle expirée - Timeout du serveur amont (upstream).",
        "505": "Version HTTP non supportée.",
        "507": "Espace de stockage insuffisant.",
        "508": "Boucle de redirection détectée.",
        # ── Codes Cloudflare (5xx étendus) ───────────────────────────────────
        "520": "Erreur inconnue Cloudflare - Le serveur d'origine a retourné une réponse inattendue.",
        "521": "Serveur web hors ligne (Cloudflare) - Connexion refusée par le serveur d'origine.",
        "522": "Timeout de connexion (Cloudflare) - Le serveur d'origine ne répond pas.",
        "523": "Origine inaccessible (Cloudflare) - Cloudflare ne peut pas joindre le serveur.",
        "524": "Timeout Cloudflare - Le serveur d'origine a mis trop de temps à répondre.",
        "525": "Handshake SSL échoué (Cloudflare) - Problème de négociation TLS avec l'origine.",
        "526": "Certificat SSL invalide (Cloudflare) - Certificat non valide côté origine.",
        "527": "Railgun Cloudflare - Erreur de connexion Railgun.",
        "530": "Site suspendu (Cloudflare) - Le site a été désactivé.",
    }
    code = str(status_code)
    if code in explanations:
        return explanations[code]
    # Famille non listée explicitement
    famille = {
        "1": "Réponse informative (1xx).",
        "2": "Succès (2xx) — code non standard.",
        "3": "Redirection (3xx) — code non standard.",
        "4": "Erreur client (4xx) — code non standard.",
        "5": "Erreur serveur (5xx) — code non standard.",
    }
    return famille.get(code[0], f"Code HTTP {code} — statut non documenté.")


import ipaddress as _ipaddress

def _detect_resina_or_intranet(domain: str, error: Exception) -> str:
    """
    Tente de déterminer si un site inaccessible est restreint au réseau
    RESINA (intranet gouvernemental BF) ou à un réseau privé.

    Retourne un message d'observation dédié, ou chaîne vide si non déterminable.
    """
    try:
        ip_str = socket.gethostbyname(domain)
        ip     = _ipaddress.ip_address(ip_str)

        # Cas 1 : IP privée RFC1918 → intranet local certain
        if ip.is_private or ip.is_loopback or ip.is_link_local:
            return (
                f"🔒 INTRANET PRIVÉ : Le domaine résout vers une adresse IP privée "
                f"({ip_str}). Ce site est uniquement accessible depuis le réseau "
                f"local ou le réseau intranet de l'organisation."
            )

        # Cas 2 : Connexion refusée (port fermé) sur IP publique
        error_str = type(error).__name__ + str(error)
        is_refused = "ConnectionRefused" in error_str or "refused" in error_str.lower()
        is_timeout = "Timeout" in error_str or "timed out" in error_str.lower()

        # Domaines gouvernementaux .gov.bf connus pour RESINA
        is_gov_bf = domain.endswith(".gov.bf") or domain.endswith(".bf")

        if is_refused and is_gov_bf:
            return (
                f"🔒 RÉSEAU RESINA PROBABLE : Connexion refusée ({ip_str}). "
                f"Ce site gouvernemental (.gov.bf) semble accessible uniquement "
                f"via le réseau intranet administratif RESINA géré par l'ANPTIC/ONATEL. "
                f"Inaccessible depuis Internet public."
            )

        if is_timeout and is_gov_bf:
            return (
                f"🔒 RÉSEAU RESINA PROBABLE : Timeout de connexion ({ip_str}). "
                f"Ce site gouvernemental (.gov.bf) est potentiellement restreint "
                f"au réseau RESINA (intranet administratif BF). "
                f"Vérifier depuis le réseau RESINA/ONATEL."
            )

        # Cas général : timeout ou refus sur n'importe quel domaine
        if is_refused:
            return (
                f"Connexion refusée par le serveur ({ip_str}). "
                f"Le port 80/443 est fermé ou filtré par un pare-feu. "
                f"Le site est peut-être restreint à un réseau intranet spécifique."
            )

    except socket.gaierror:
        # DNS ne résout pas → domaine inexistant ou DNS interne uniquement
        if domain.endswith(".gov.bf"):
            return (
                "🔒 DNS NON PUBLIC : Le domaine ne se résout pas depuis les DNS publics. "
                "Il est probablement enregistré uniquement dans le DNS interne RESINA. "
                "Accessible uniquement depuis le réseau administratif gouvernemental BF."
            )
        return "Domaine non résolu — DNS inexistant ou domaine expiré."

    except Exception:
        pass

    return ""


def detect_advanced_technologies(url: str, response: requests.Response) -> str:
    """
    Détection native de technologies sans dépendance externe.
    Analyse les headers HTTP et le corps HTML pour identifier :
    CMS, frameworks, serveurs, langages, CDN, outils analytiques.
    """
    found = []
    headers = {k.lower(): v for k, v in response.headers.items()}
    html    = response.text[:50_000].lower()   # 50 ko suffisent, évite les gros DOM

    # ── Serveur web (header Server) ──────────────────────────────────────────
    server = headers.get("server", "")
    if server:
        if   "nginx"   in server.lower(): found.append("Nginx")
        elif "apache"  in server.lower(): found.append("Apache")
        elif "iis"     in server.lower(): found.append("Microsoft IIS")
        elif "litespeed" in server.lower(): found.append("LiteSpeed")
        elif "caddy"   in server.lower(): found.append("Caddy")
        elif server.strip():              found.append(f"Serveur: {server.split()[0]}")

    # ── Langage / runtime (header X-Powered-By) ──────────────────────────────
    powered = headers.get("x-powered-by", "")
    if powered:
        if   "php"        in powered.lower(): found.append(f"PHP ({powered.strip()})")
        elif "asp.net"    in powered.lower(): found.append("ASP.NET")
        elif "express"    in powered.lower(): found.append("Express.js")
        elif "next.js"    in powered.lower(): found.append("Next.js")
        elif powered.strip():                 found.append(powered.strip())

    # ── CDN / Proxy (headers spécifiques) ────────────────────────────────────
    if "cf-ray"          in headers: found.append("Cloudflare")
    if "x-amz-cf-id"     in headers: found.append("AWS CloudFront")
    if "x-cache"         in headers and "varnish" in headers.get("x-cache","").lower():
        found.append("Varnish Cache")
    if "x-sucuri-id"     in headers: found.append("Sucuri WAF")
    if "x-fw-hash"       in headers: found.append("Flywheel")

    # ── CMS (HTML) ───────────────────────────────────────────────────────────
    if "/wp-content/"    in html or "/wp-includes/" in html: found.append("WordPress")
    if "joomla"          in html or "/components/com_" in html: found.append("Joomla")
    if "drupal"          in html or "drupal.settings" in html: found.append("Drupal")
    if "sites/default/files" in html:                          found.append("Drupal")
    if "/typo3/"         in html:                              found.append("TYPO3")
    if "spip"            in html:                              found.append("SPIP")
    if "prestashop"      in html:                              found.append("PrestaShop")
    if "shopify"         in html:                              found.append("Shopify")
    if "wix.com"         in html:                              found.append("Wix")
    if "squarespace"     in html:                              found.append("Squarespace")

    # ── Frameworks JS (HTML) ─────────────────────────────────────────────────
    if "react"           in html or "react-dom" in html:      found.append("React")
    if "vue.js"          in html or "vue.min.js" in html:     found.append("Vue.js")
    if "angular"         in html:                              found.append("Angular")
    if "jquery"          in html:                              found.append("jQuery")
    if "bootstrap"       in html:                              found.append("Bootstrap")
    if "next.js"         in html or "__next" in html:         found.append("Next.js")
    if "nuxt"            in html:                              found.append("Nuxt.js")

    # ── Analytics / Tag managers ─────────────────────────────────────────────
    if "google-analytics.com" in html or "gtag(" in html:    found.append("Google Analytics")
    if "googletagmanager.com" in html:                        found.append("Google Tag Manager")
    if "matomo"          in html or "piwik"  in html:         found.append("Matomo")
    if "facebook.net/fr_fr/fbevents" in html:                 found.append("Facebook Pixel")

    # ── Dédoublonnage et formatage ───────────────────────────────────────────
    seen  = set()
    uniq  = [x for x in found if not (x in seen or seen.add(x))]

    if uniq:
        return f"Technologies détectées : {', '.join(uniq)}."
    return "Aucune technologie majeure détectée."


def analyze_domain_infrastructure(domain: str):
    """
    Analyse l'hébergeur, le DNSSEC et la date de création du domaine.

    CORRECTIF : l'ancienne logique if/elif manquait de nombreux opérateurs
    (Orange BF, Telecel, MTN…) et le fallback split('-')[0] tronquait
    les noms commençant par 'AS-' (ex: 'AS-ORANGE-BF' → 'AS' au lieu
    d'Orange). Remplacé par une table de correspondance exhaustive.
    """
    hosting       = "non défini"
    dnssec_status = "Non activé"
    creation_date = "Non définie"

    if not domain:
        return hosting, dnssec_status, creation_date

    # ── Hébergeur via RDAP ───────────────────────────────────────────────────
    try:
        ip       = socket.gethostbyname(domain)
        rdap_res = _cached_rdap(ip)
        asn_desc = rdap_res.get("asn_description", "").upper()

        # Table de correspondance ordonnée par priorité.
        # Chaque entrée : (sous-chaîne cherchée dans asn_desc, label affiché)
        # Les entrées plus spécifiques (ex: 'AS-ORANGE-BF') passent AVANT
        # les termes génériques (ex: 'ORANGE') pour éviter les faux positifs.
        _PROVIDERS = [
            # ── Burkina Faso ─────────────────────────────────────────────
            ("AS-ORANGE-BF",      "Orange Burkina Faso"),
            ("ORANGE-BF",         "Orange Burkina Faso"),
            ("ONATEL",            "ONATEL"),
            ("ANPTIC",            "ANPTIC"),
            ("TELECEL-FASO",      "Telecel Faso"),
            ("TELECEL",           "Telecel Faso"),
            ("FASOCELL",          "Fasocell"),
            ("LIPTINFOR",         "Liptinfor"),
            ("MTN-BFASO",         "MTN Burkina Faso"),
            ("CANAL-BF",          "Canal BF"),
            ("ZOOMTEL",           "Zoomtel"),
            ("AZIMUT-BF",         "Azimut BF"),
            ("OPEN-BF",           "Open BF"),
            ("LAFIABAT",          "Lafiabat"),
            # ── Afrique de l'Ouest / sous-région ────────────────────────
            ("SONATEL",           "Orange Sénégal (Sonatel)"),
            ("CAMTEL",            "Camtel (Cameroun)"),
            ("MTN",               "MTN"),
            ("MOOV",              "Moov Africa"),
            ("AFRIPA",            "Afripa Telecom"),
            ("CELTEL",            "Celtel/Airtel"),
            ("AIRTEL",            "Airtel"),
            ("ORANGE",            "Orange"),
            ("AFRICELL",          "Africell"),
            ("ISOCEL",            "Isocel Telecom"),
            # ── Hébergeurs internationaux ────────────────────────────────
            ("OVH",               "OVH"),
            ("CLOUDFLARENET",     "Cloudflare"),
            ("CLOUDFLARE",        "Cloudflare"),
            ("AMAZON",            "AWS"),
            ("AWS",               "AWS"),
            ("MICROSOFT",         "Azure"),
            ("AZURE",             "Azure"),
            ("GOOGLE",            "Google Cloud"),
            ("HETZNER",           "Hetzner"),
            ("DIGITALOCEAN",      "DigitalOcean"),
            ("LINODE",            "Linode/Akamai"),
            ("AKAMAI",            "Akamai"),
            ("SCALEWAY",          "Scaleway"),
            ("GANDI",             "Gandi"),
            ("IONOS",             "IONOS"),
            ("HOSTINGER",         "Hostinger"),
            ("CONTABO",           "Contabo"),
        ]

        hosting = "non défini"
        for keyword, label in _PROVIDERS:
            if keyword in asn_desc:
                hosting = label
                break
        else:
            # ── Fallback exhaustif : on cherche le meilleur nom disponible ──
            # Priorité 1 : asn_description brut (avant la virgule du code pays)
            #   ex: "AS-EXAMPLE-BF, BF" → "AS-EXAMPLE-BF"
            raw_asn = rdap_res.get("asn_description", "").split(",")[0].strip()

            # Priorité 2 : network.name (souvent plus lisible que asn_description)
            #   ex: {"name": "EXAMPLE-NET"} → "EXAMPLE-NET"
            net_name = rdap_res.get("network", {}).get("name", "").strip()

            # Priorité 3 : nom de l'entité de type "registrant" ou "abuse"
            #   Les entités RDAP portent parfois la raison sociale complète
            entity_name = ""
            for ent in rdap_res.get("entities", []):
                contact = ent.get("contact", {})
                name    = contact.get("name", "")
                roles   = ent.get("roles", [])
                # On préfère l'entité registrant/technical, pas abuse/noc
                if name and any(r in roles for r in ("registrant", "technical", "administrative")):
                    entity_name = name
                    break
            if not entity_name:
                # Dernier recours : première entité avec un nom
                for ent in rdap_res.get("entities", []):
                    name = ent.get("contact", {}).get("name", "")
                    if name:
                        entity_name = name
                        break

            # On choisit la valeur la plus informative
            hosting = raw_asn or net_name or entity_name or "non défini"

    except Exception:
        hosting = "Inaccessible"

    # ── DNSSEC ───────────────────────────────────────────────────────────────
    try:
        answers = dns.resolver.resolve(domain, 'RRSIG', raise_on_no_answer=True)
        if len(answers) > 0:
            dnssec_status = "Activé"
    except Exception:
        dnssec_status = "Non activé"

    # ── Date de création (whois) ─────────────────────────────────────────────
    # Priorité : creation_date → updated_date → expiration_date
    # Certains registrars masquent la création mais exposent la mise à jour.
    try:
        info = _cached_whois(domain)
        if info:
            date_val = None
            date_label = ""

            # Priorité 1 : date de création (mise en ligne réelle)
            raw = info.creation_date
            if raw:
                date_val  = raw[0] if isinstance(raw, list) else raw
                date_label = ""   # label vide = c'est bien la date de création

            # Priorité 2 : date de dernière mise à jour whois
            if not date_val:
                raw = info.updated_date
                if raw:
                    date_val  = raw[0] if isinstance(raw, list) else raw
                    date_label = "MAJ: "  # préfixe pour distinguer

            # Priorité 3 : date d'expiration (au moins on sait que le domaine existe)
            if not date_val:
                raw = info.expiration_date
                if raw:
                    date_val  = raw[0] if isinstance(raw, list) else raw
                    date_label = "Exp: "

            if date_val:
                creation_date = date_label + date_val.strftime("%d/%m/%Y")
            else:
                creation_date = "Masqué / Erreur"
    except Exception:
        creation_date = "Masqué / Erreur"

    return hosting, dnssec_status, creation_date


# ──────────────────────────────────────────────────────────────────────────────
# Stratégie de requête HTTP en 3 passes  (FIX A + FIX B)
# ──────────────────────────────────────────────────────────────────────────────
# FIX B – Timeouts adaptés aux serveurs gouvernementaux lents
_TIMEOUT_CONNECT = 6.0   # secondes pour établir la connexion TCP
_TIMEOUT_READ    = 12.0  # secondes pour recevoir la réponse


def _fetch_with_fallback(url_https: str):
    """
    Tente d'atteindre un site en 3 passes successives.

    Retourne (response, url_used, mode) où mode est :
      "https_ok"      – HTTPS valide
      "https_nossl"   – HTTPS avec verify=False (cert invalide mais site UP)
      "http_only"     – HTTP seulement (pas de TLS)
      None            – site inaccessible (Exception propagée)

    FIX A : on tente verify=False avant de conclure que le site est DOWN.
            Beaucoup de sites gouvernementaux .bf ont des certificats
            auto-signés ou signés par une CA interne non reconnue.
    """
    session = _get_session()
    url_http = url_https.replace("https://", "http://", 1)

    # Passe 1 : HTTPS standard (certificat vérifié)
    try:
        resp = session.get(
            url_https,
            timeout=(_TIMEOUT_CONNECT, _TIMEOUT_READ),
            allow_redirects=True,
            verify=True,
        )
        # Si la redirection finale atterrit en HTTP → on note http_only
        mode = "http_only" if resp.url.startswith("http://") else "https_ok"
        return resp, resp.url, mode
    except requests.exceptions.SSLError:
        pass  # → Passe 2
    except (requests.exceptions.ConnectionError,
            requests.exceptions.Timeout):
        pass  # → Passe 3 directement (inutile de réessayer en HTTPS)

    # Passe 2 : HTTPS sans vérification du certificat
    #           (certificat auto-signé, CA interne, cert expiré récemment…)
    try:
        resp = session.get(
            url_https,
            timeout=(_TIMEOUT_CONNECT, _TIMEOUT_READ),
            allow_redirects=True,
            verify=False,      # FIX A
        )
        mode = "http_only" if resp.url.startswith("http://") else "https_nossl"
        return resp, resp.url, mode
    except requests.exceptions.SSLError:
        pass  # cert vraiment inutilisable → Passe 3
    except (requests.exceptions.ConnectionError,
            requests.exceptions.Timeout):
        pass  # → Passe 3

    # Passe 3 : HTTP pur (port 80)
    resp = session.get(
        url_http,
        timeout=(_TIMEOUT_CONNECT, _TIMEOUT_READ),
        allow_redirects=True,
        verify=False,
    )
    return resp, resp.url, "http_only"


# ──────────────────────────────────────────────────────────────────────────────
# Analyse principale
# ──────────────────────────────────────────────────────────────────────────────

def check_website_complete(url: str) -> dict:
    """Analyse complète d'un site web."""
    if not url.startswith(("http://", "https://")):
        url_to_request = "https://" + url
    else:
        url_to_request = (
            url if url.startswith("https://")
            else url.replace("http://", "https://", 1)
        )

    domain    = urlparse(url_to_request).netloc
    scan_time = datetime.datetime.now().strftime("%d/%m/%Y")

    results = {
        "url":               url,
        "disponible":        "NON",
        "statut_http":       "Inconnu",
        "certificat":        "INCONNU",
        "hosting":           "non défini",
        "date_mise_en_ligne":"Non définie",
        "last_modified":     "Inconnue",
        "dnssec":            "Non activé",
        "observation":       "",
        "date_scan":         scan_time,
        "is_http_only":      False,
    }

    # 1. Infrastructure
    hosting, dnssec, creation = analyze_domain_infrastructure(domain)
    results.update(hosting=hosting, dnssec=dnssec, date_mise_en_ligne=creation)

    # 2. Requête HTTP (stratégie 3 passes)
    try:
        response, final_url, mode = _fetch_with_fallback(url_to_request)
        results["statut_http"] = str(response.status_code)
        results["is_http_only"] = (mode == "http_only")

        # ── Disponibilité selon code HTTP ────────────────────────────────────
        code = response.status_code
        if code == 200:
            results["disponible"] = "OUI (Fonctionnel)"
        elif code in (301, 302, 303, 307, 308):
            results["disponible"] = f"OUI (Redirection {code})"
        elif code in (401, 403):
            results["disponible"] = f"OUI (Accès restreint {code})"
        elif code == 404:
            results["disponible"] = "OUI - Page 404 (Contenu introuvable)"
        elif code == 429:
            results["disponible"] = "OUI (Limite de débit - 429)"
        elif code in (500, 501, 502, 503, 504, 505, 520, 521, 522, 523, 524, 525, 526, 530):
            results["disponible"] = f"OUI - Erreur serveur ({code})"
        else:
            results["disponible"] = f"OUI - Code HTTP {code}"

        # ── Construction de l'Observation ────────────────────────────────────
        explanation = get_http_status_explanation(code)
        alerts      = []

        # Page Apache par défaut
        if code == 200 and _is_apache_default_page(response.text):
            results["disponible"] = "OUI (Page Apache par défaut - non configuré)"
            alerts.append(
                "⚠️ SERVEUR NON CONFIGURÉ : Le site affiche la page Apache par défaut "
                "(«It works!»). Aucun contenu réel n'est déployé sur ce serveur."
            )

        # Certificat invalide mais site accessible
        if mode == "https_nossl":
            alerts.append(
                "⚠️ CERTIFICAT SSL INVALIDE : Le site est accessible mais son certificat "
                "n'est pas reconnu par les autorités de certification standard "
                "(auto-signé ou CA interne). Les navigateurs afficheront un avertissement."
            )

        # Site HTTP uniquement
        if mode == "http_only":
            alerts.append(
                "⚠️ HTTP NON SÉCURISÉ : Ce site utilise uniquement le protocole HTTP "
                "(les échanges ne sont pas chiffrés). Aucun certificat SSL/TLS actif."
            )

        tech_info = detect_advanced_technologies(final_url, response)
        obs_parts = [explanation] + alerts + [tech_info]
        results["observation"] = " | ".join(obs_parts)

        results["last_modified"] = parse_http_date(
            response.headers.get("Last-Modified", "")
        )

    except Exception as exc:
        # ── Diagnostic RESINA / intranet avant de conclure inaccessible ──────
        resina_msg = _detect_resina_or_intranet(domain, exc)

        if resina_msg:
            results["disponible"]  = "NON ACCESSIBLE DEPUIS INTERNET"
            results["statut_http"] = "Filtré/Intranet"
            results["observation"] = resina_msg
        else:
            results["disponible"]  = "NON (Inaccessible / Timeout)"
            results["statut_http"] = "Timeout/Error"
            results["observation"] = (
                "Impossible de joindre le serveur après 3 tentatives "
                "(HTTPS valide → HTTPS sans vérif. SSL → HTTP). "
                "Cause probable : timeout réseau, blocage pare-feu ou domaine inexistant."
            )

    # 3. Inspection TLS (seulement si le site supporte HTTPS)
    if not results["is_http_only"] and results["statut_http"] not in ("SSL_Error", "Timeout/Error", "Inconnu"):
        try:
            context = ssl.create_default_context()
            # FIX A : on tente d'abord avec verify, puis sans (pour cert auto-signé)
            try:
                with socket.create_connection((domain, 443), timeout=4) as sock:
                    with context.wrap_socket(sock, server_hostname=domain) as ssock:
                        tls_version = ssock.version().replace("TLSv", "TLS ")
                        cert        = ssock.getpeercert()
                        expire_date = datetime.datetime.strptime(
                            cert["notAfter"], "%b %d %H:%M:%S %Y %Z"
                        )
                        jours = (expire_date - datetime.datetime.utcnow()).days
                        results["certificat"] = f"{tls_version} ({jours} j restants)"
            except ssl.SSLCertVerificationError:
                # Certificat invalide : on récupère quand même les infos TLS
                ctx_noverif = ssl.create_default_context()
                ctx_noverif.check_hostname = False
                ctx_noverif.verify_mode    = ssl.CERT_NONE
                with socket.create_connection((domain, 443), timeout=4) as sock:
                    with ctx_noverif.wrap_socket(sock, server_hostname=domain) as ssock:
                        tls_version = ssock.version().replace("TLSv", "TLS ")
                        results["certificat"] = f"{tls_version} (cert invalide/auto-signé)"
        except Exception:
            results["certificat"] = "INCONNU"
    else:
        results["certificat"] = "NON APPLICABLE (HTTP)"

    return results


# ──────────────────────────────────────────────────────────────────────────────
# Orchestration Excel
# ──────────────────────────────────────────────────────────────────────────────

def process_excel_concurrent(filepath: str, max_workers: int = 30):
    start_time = time.time()

    #// data_only=True : openpyxl retourne la VALEUR CALCULÉE des formules,
    #// pas la formule brute. INDISPENSABLE pour les cellules avec
    #// =SI(ESTNUM(CHERCHE(...));"";<numéro>) en colonne A.
    #// Sans data_only=True, col_a vaudrait "=SI(...)" (chaîne) → jamais traité.
    #wb = openpyxl.load_workbook(filepath, data_only=True)
    #//ws = wb.active # première feuille active (généralement la seule)
    #ws = wb.worksheets[1] # deuxième feuille (index 1) pour les fichiers multi-feuilles 
    #print(f"Traitement démarré sur la feuille : {ws.title}")

 # 1. Lecture de la Feuille 2 du fichier d'origine
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if len(wb.sheetnames) < 2:
        print("⚠️-Erreur : Le fichier ne contient pas de deuxième feuille.\n")
        print("▶️-Traitement démarré sur la feuille active.\n")
        ws = wb.active # première feuille active (généralement la seule)
    else:    
        # Sélection de la deuxième feuille (index 1)
        ws = wb.worksheets[1]
        print(f"Traitement démarré sur la feuille : {ws.title}")

    tasks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        col_a = row[0].value
        col_b = row[1].value if len(row) > 1 else None

        # La formule =SI(ESTNUM(CHERCHE("▶";Bx));"";SOMMEPROD(...)) retourne :
        #   → un int ou float (ex: 1.0, 2.0…) quand la ligne est une donnée
        #   → ""  (chaîne vide)               quand la ligne est un groupe ▶
        #   → None                             quand la cellule est vraiment vide
        #
        # Condition de traitement :
        #   A doit être un nombre réel (int ou float, pas bool, pas chaîne)
        #   B doit être une chaîne non vide (l'URL)
        a_est_nombre = (
            isinstance(col_a, (int, float))
            and not isinstance(col_a, bool)
        )
        b_est_chaine = (
            isinstance(col_b, str)
            and col_b.strip() != ""
        )

        if not a_est_nombre or not b_est_chaine:
            # En-tête, ligne de catégorie (▶), ligne vide → on ignore
            # On n'affiche le message que si au moins une cellule a une valeur
            if col_a is not None or col_b is not None:
                print(f"  [Ligne {row_idx}] Ignorée — A={repr(col_a)} | B={repr(str(col_b)[:60])}")
            continue

        tasks.append((row_idx, col_b.strip()))

    total = len(tasks)
    print(f"Lancement de l'analyse sur {total} sites ({max_workers} threads)\n")

    completed = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(check_website_complete, url): (row_idx, url)
            for row_idx, url in tasks
        }

        for future in as_completed(future_map):
            row_idx, url = future_map[future]
            completed += 1
            try:
                data = future.result()
                ws.cell(row=row_idx, column=5,  value=data["disponible"])
                ws.cell(row=row_idx, column=6,  value=data["certificat"])
                ws.cell(row=row_idx, column=7,  value=data["hosting"])
                ws.cell(row=row_idx, column=8,  value=data["date_mise_en_ligne"])
                ws.cell(row=row_idx, column=9,  value=data["last_modified"])
                ws.cell(row=row_idx, column=10, value=data["dnssec"])
                ws.cell(row=row_idx, column=11, value=data["observation"])
                ws.cell(row=row_idx, column=12, value=data["date_scan"])

                print(f"[{completed:>4}/{total}] {url} → HTTP {data['statut_http']}")
                print(f"  Disponible : {data['disponible']}")
                print(f"  SSL/TLS    : {data['certificat']}")
                print(f"  Hébergeur  : {data['hosting']}")
                print(f"  Création   : {data['date_mise_en_ligne']}")
                print(f"  Modif.     : {data['last_modified']}")
                print(f"  DNSSEC     : {data['dnssec']}")
                print(f"  Observation: {data['observation']}")
                print("─" * 76)
            except Exception as exc:
                print(f"  ERREUR ligne {row_idx} ({url}) : {exc}")

# 1. Définir le nom du dossier et le nom du fichier
    nom_dossier = "resultats"

    horodatage      = datetime.datetime.now().strftime("%d-%m-%Y_%Hh-%M-%S")
    output_filename = f"res_sites_20{horodatage}.xlsx"
    #output_filename = f"zak_resultats_sites_{horodatage}.xlsx"
    #output_filename = f"rayhana_resultats_sites_{horodatage}.xlsx"
    #output_filename = f"dri_resultats_sites_{horodatage}.xlsx"
    
    # 2. Créer le dossier automatiquement s'il n'existe pas encore
    os.makedirs(nom_dossier, exist_ok=True)

    # 3. Combiner le dossier et le fichier pour obtenir le chemin complet
    chemin_complet = os.path.join(nom_dossier, output_filename)
    wb.save(chemin_complet)

    elapsed        = time.time() - start_time
    minutes, secs  = divmod(int(elapsed), 60)
    print(f"\n✅ Traitement terminé → Sauvegardé dans : {chemin_complet}")
    print(f"⏱  Temps total : {minutes:02d}m {secs:02d}s\n")


if __name__ == "__main__":
    fichier_liste = sys.argv[1] if len(sys.argv) > 1 else "sites_20.xlsx"
    #fichier_liste = sys.argv[1] if len(sys.argv) > 1 else "zak_924_a_972_plateformes_burkina_faso.xlsx"
    #fichier_liste = sys.argv[1] if len(sys.argv) > 1 else "rayhana_973_a_1021_plateformes_burkina_faso.xlsx"
    #fichier_liste = sys.argv[1] if len(sys.argv) > 1 else "equipe_e1_sites_sawadogo_idrissa_071838.xlsx"

    workers       = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    process_excel_concurrent(fichier_liste, max_workers=workers)